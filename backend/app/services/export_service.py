"""
TalentMatch AI - Export Service
Generates CSV and Excel exports of candidate rankings
"""

import io
import logging
from typing import List, Dict, Any
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportService:
    """Handles CSV and Excel export of candidate data."""

    def generate_csv(self, candidates: List[Dict[str, Any]], jd_title: str) -> bytes:
        """
        Generate CSV export of candidates.

        Returns:
            CSV content as bytes
        """
        import csv

        output = io.StringIO()
        fieldnames = [
            "Rank", "Name", "Email", "Phone", "Location",
            "Total Score", "Skills Score", "Experience Score",
            "Education Score", "Keyword Score", "Project Score",
            "Years of Experience", "Education Level",
            "Matched Skills", "Missing Skills",
            "Recommendation", "Resume File", "Upload Date"
        ]

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()

        for candidate in candidates:
            analysis = candidate.get("analysis_result") or {}
            skills = [s.get("skill", "") for s in candidate.get("skills", [])]

            writer.writerow({
                "Rank": analysis.get("rank", "N/A"),
                "Name": candidate.get("name", ""),
                "Email": candidate.get("email", ""),
                "Phone": candidate.get("phone", ""),
                "Location": candidate.get("location", ""),
                "Total Score": f"{analysis.get('total_score', 0):.1f}",
                "Skills Score": f"{analysis.get('skills_score', 0):.1f}/40",
                "Experience Score": f"{analysis.get('experience_score', 0):.1f}/25",
                "Education Score": f"{analysis.get('education_score', 0):.1f}/15",
                "Keyword Score": f"{analysis.get('keyword_score', 0):.1f}/10",
                "Project Score": f"{analysis.get('project_score', 0):.1f}/10",
                "Years of Experience": candidate.get("years_of_experience", 0),
                "Education Level": candidate.get("education_level", ""),
                "Matched Skills": ", ".join(analysis.get("matched_skills", [])),
                "Missing Skills": ", ".join(analysis.get("missing_skills", [])),
                "Recommendation": analysis.get("recommendation", ""),
                "Resume File": candidate.get("resume_filename", ""),
                "Upload Date": str(candidate.get("created_at", ""))[:10],
            })

        return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility

    def generate_excel(self, candidates: List[Dict[str, Any]], jd_title: str) -> bytes:
        """
        Generate Excel (.xlsx) export with formatting.

        Returns:
            Excel file content as bytes
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side,
                GradientFill
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl not installed. Falling back to CSV.")
            return self.generate_csv(candidates, jd_title)

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidate Rankings"

        # ─────────────────────────────────────────
        # Styles
        # ─────────────────────────────────────────
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1E293B")  # Dark slate
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        title_font = Font(bold=True, size=14, color="1E293B")

        score_fills = {
            "excellent": PatternFill("solid", fgColor="DCFCE7"),   # Green
            "good": PatternFill("solid", fgColor="DBEAFE"),         # Blue
            "average": PatternFill("solid", fgColor="FEF9C3"),      # Yellow
            "poor": PatternFill("solid", fgColor="FEE2E2"),         # Red
        }

        def get_score_fill(score: float):
            if score >= 85:
                return score_fills["excellent"]
            elif score >= 70:
                return score_fills["good"]
            elif score >= 50:
                return score_fills["average"]
            return score_fills["poor"]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ─────────────────────────────────────────
        # Title Row
        # ─────────────────────────────────────────
        ws.merge_cells("A1:R1")
        ws["A1"] = f"TalentMatch AI – Candidate Rankings | {jd_title}"
        ws["A1"].font = title_font
        ws["A1"].fill = PatternFill("solid", fgColor="F0F9FF")
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:R2")
        ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total Candidates: {len(candidates)}"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 18

        # ─────────────────────────────────────────
        # Headers
        # ─────────────────────────────────────────
        headers = [
            "Rank", "Name", "Email", "Phone", "Location",
            "Total Score", "Skills\n(/40)", "Experience\n(/25)",
            "Education\n(/15)", "Keywords\n(/10)", "Projects\n(/10)",
            "Years Exp", "Education Level",
            "Matched Skills", "Missing Skills",
            "Recommendation", "Resume File", "Upload Date"
        ]

        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[header_row].height = 35

        # ─────────────────────────────────────────
        # Data Rows
        # ─────────────────────────────────────────
        for row_idx, candidate in enumerate(candidates, start=header_row + 1):
            analysis = candidate.get("analysis_result") or {}
            score = analysis.get("total_score", 0)
            score_fill = get_score_fill(score)

            row_data = [
                analysis.get("rank", "N/A"),
                candidate.get("name", ""),
                candidate.get("email", ""),
                candidate.get("phone", ""),
                candidate.get("location", ""),
                round(score, 1),
                round(analysis.get("skills_score", 0), 1),
                round(analysis.get("experience_score", 0), 1),
                round(analysis.get("education_score", 0), 1),
                round(analysis.get("keyword_score", 0), 1),
                round(analysis.get("project_score", 0), 1),
                candidate.get("years_of_experience", 0),
                candidate.get("education_level", ""),
                ", ".join(analysis.get("matched_skills", [])[:5]),
                ", ".join(analysis.get("missing_skills", [])[:5]),
                analysis.get("recommendation", ""),
                candidate.get("resume_filename", ""),
                str(candidate.get("created_at", ""))[:10],
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col in [14, 15]))

                # Apply score color to score columns
                if col == 6 and isinstance(value, (int, float)):
                    cell.fill = score_fill
                    cell.font = Font(bold=True)

            ws.row_dimensions[row_idx].height = 20

        # ─────────────────────────────────────────
        # Column Widths
        # ─────────────────────────────────────────
        col_widths = [8, 22, 28, 16, 18, 12, 10, 12, 10, 10, 10, 10, 18, 35, 35, 20, 25, 12]
        for col, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header rows
        ws.freeze_panes = f"A{header_row + 1}"

        # Auto-filter
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row}"

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()


# Singleton instance
export_service = ExportService()
